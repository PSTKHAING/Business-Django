from django.shortcuts import render,redirect
from django.contrib import messages
from django.contrib.auth.hashers import check_password
from django.contrib.auth import login,logout
from django.conf import settings
from django.contrib.auth.decorators import login_required
from authentication.decorators import role_permission_required
from django.contrib.auth.models import Permission
from authentication.models import *
import openpyxl
from django.http import HttpResponse
from django.db.models import Q
from django.core.paginator import Paginator
import urllib.parse
from datetime import datetime
from django.db import IntegrityError

@login_required(login_url=settings.LOGIN_URL)
# @role_permission_required('authentication.add_usermodel')
def UserCreate(request):
    if request.method == "POST":
        username = request.POST.get('username')
        email = request.POST.get("email")
        profile = request.FILES.get('profile')
        is_active = request.POST.get('is_active') == 'on'
        is_staff = request.POST.get('is_staff') == 'on'
        is_superuser = request.POST.get('is_superuser') == 'on'
        role_id = request.POST.get('role')
        
        if UserModel.objects.filter(email=email).exists():
            messages.error(request, "Email has already been used!")
            return redirect('/user/list/')
        
        password = request.POST.get('password')
        password_confirm = request.POST.get('passwordconfirm')
        if password == password_confirm:
            user = UserModel.objects.create_user(
                username=username,
                email=email,
                profile=profile,
                password=password,
                is_active=is_active,
                is_staff=is_staff,
                is_superuser=is_superuser,
                role_id=role_id,
                created_by =request.user
            )
            user.save()
            messages.success(request, "Account was created for " + username)
            return redirect(f'/user/list/')
        else:
            messages.error(request, "Password does not match! Please check your password again!")
            return redirect('/user/list/')

@login_required(login_url=settings.LOGIN_URL)
# @role_permission_required('authentication.change_usermodel')
@login_required(login_url=settings.LOGIN_URL)
def UserUpdate(request, pk):
    user = UserModel.objects.get(id=pk)
    search_term = request.GET.get('search', '')
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')
    pick_date = request.GET.get('pick_date', '')
    page_number = request.GET.get('page', 1)
    entries = request.GET.get('entries', 5)

    if from_date:
        from_date = datetime.strptime(from_date, '%d %B, %Y')
    if to_date:
        to_date = datetime.strptime(to_date, '%d %B, %Y')
    if pick_date:
        pick_date = datetime.strptime(pick_date, '%d %B, %Y')

    if request.method == 'POST':
        user.email = request.POST.get('email')
        user.username = request.POST.get('username')
        if request.FILES.get('profile'):
            user.profile.delete()
            user.profile = request.FILES.get('profile')
        user.is_active = request.POST.get('is_active') == 'on'
        user.is_staff = request.POST.get('is_staff') == 'on'
        user.is_superuser = request.POST.get('is_superuser') == 'on'
        user.role_id = request.POST.get('role')
        if request.POST.get('password'):
            user.set_password(request.POST.get('password'))
        user.updated_by = request.user
        user.save()
        messages.success(request, "User was updated!")

        return redirect(f'/user/list/?page={page_number}&search={search_term}&entries={entries}&from_date={from_date}&to_date={to_date}&pick_date={pick_date}')

@login_required(login_url=settings.LOGIN_URL)
# @role_permission_required('authentication.delete_usermodel')
def UserDelete(request, pk):
    user = UserModel.objects.get(id=pk)
    user.soft_delete(request.user)
    messages.success(request, "User has been deleted!")
    return redirect('/user/list/')

@login_required(login_url=settings.LOGIN_URL)
# @role_permission_required('authentication.delete_usermodel')
def UserDeletePermanent(request, pk):
    user = UserModel.objects.get(id=pk)
    if user.profile:
        user.profile.delete()
    user.delete()
    messages.success(request, "User has been deleted!")
    return redirect('/user/trash/list/')

@login_required(login_url=settings.LOGIN_URL)
def UserDeleteSelected(request):
    if request.method == 'POST':
        selected_user_ids = request.POST.getlist('selected_users')

        if selected_user_ids:
            try:
                users = UserModel.objects.filter(id__in=selected_user_ids)
                for user in users:
                    user.soft_delete(request.user)
                messages.success(request, f"{len(users)} user(s) deleted successfully.")
                return redirect('/user/list/')
            except Exception as e:
                messages.error(request, f"An error occurred while deleting users: {str(e)}")
                return redirect('/user/list/')
        else:
            messages.error(request, 'No users selected for deletion.')
            return redirect('/user/list/')

@login_required(login_url=settings.LOGIN_URL)
def UserTrashManageSelected(request):
    if request.method == 'POST':
        selected_user_ids = request.POST.getlist('selected_users')
        action = request.POST.get('action')

        if selected_user_ids:
            try:
                users = UserModel.objects.filter(id__in=selected_user_ids)
                if action == 'delete':
                    for user in users:
                        user.delete()
                    messages.success(request, f"{len(users)} user(s) deleted permanently.")
                elif action == 'restore':
                    for user in users:
                        user.restore()
                    messages.success(request, f"{len(users)} user(s) restored successfully.")
                else:
                    messages.error(request, 'Invalid action.')
                return redirect('/user/trash/list/')
            except Exception as e:
                messages.error(request, f"An error occurred: {str(e)}")
                return redirect('/user/trash/list/')
        else:
            messages.error(request, 'No users selected.')
            return redirect('/user/trash/list/')

@login_required(login_url=settings.LOGIN_URL)
# @role_permission_required('authentication.view_usermodel')
def UserList(request):
    search_term = request.GET.get('search', '')
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')
    pick_date = request.GET.get('pick_date', '')
    page_number = request.GET.get('page', 1)
    entries = request.GET.get('entries', 5)
    try:
        entries = int(entries)
    except ValueError:
        entries = 5
    users = UserModel.objects.filter(deleted_at = None).order_by('-created_at')
    trash_users = UserModel.objects.filter(deleted_at__isnull=False).order_by('-deleted_at')

    if search_term:
        users = users.filter(
            Q(username__icontains=search_term) |
            Q(email__icontains=search_term) |
            Q(role__name__icontains=search_term)
        )
    if from_date:
        try:
            from_date = datetime.strptime(from_date, '%d %B, %Y')
            users = users.filter(date_joined__gte=from_date)
        except ValueError:
            messages.error(request, "Invalid From Date format. Please use '4 March, 2025'.")

    if to_date:
        try:
            to_date = datetime.strptime(to_date, '%d %B, %Y')
            to_date = to_date.replace(hour=23, minute=59, second=59, microsecond=999999)
            users = users.filter(date_joined__lte=to_date)
        except ValueError:
            messages.error(request, "Invalid To Date format. Please use '4 March, 2025'.")

    if pick_date:
        try:
            pick_date = datetime.strptime(pick_date, '%d %B, %Y')
            pick_date_start = pick_date.replace(hour=0, minute=0, second=0, microsecond=0)
            pick_date_end = pick_date.replace(hour=23, minute=59, second=59, microsecond=999999)
            users = users.filter(date_joined__range=[pick_date_start, pick_date_end])
        except ValueError:
            messages.error(request, "Invalid Pick Date format. Please use '4 March, 2025'.")

    paginator = Paginator(users, entries)
    page_obj = paginator.get_page(page_number)

    roles = RoleModel.objects.all().order_by('-created_at')

    context = {
        "users": page_obj,
        "trash_users" : trash_users,
        "roles": roles,
        "search": search_term,
        "entries": entries,
        "from_date": from_date,
        "to_date": to_date,
        "pick_date": pick_date,
    }
    return render(request, 'dashboard/user_list.html', context)

def UserExportExcel(request):
    search_term = request.GET.get('search', '')
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')
    pick_date = request.GET.get('pick_date', '')

    users = UserModel.objects.all().order_by('-created_at')

    if search_term:
        users = users.filter(
            Q(username__icontains=search_term) |
            Q(email__icontains=search_term) |
            Q(role__name__icontains=search_term)
        )

    if from_date:
        try:
            from_date = datetime.strptime(from_date, '%d %B, %Y')
            users = users.filter(date_joined__gte=from_date)
        except ValueError:
            messages.error(request, "Invalid From Date format. Please use '4 March, 2025'.")

    if to_date:
        try:
            to_date = datetime.strptime(to_date, '%d %B, %Y')
            to_date = to_date.replace(hour=23, minute=59, second=59, microsecond=999999)
            users = users.filter(date_joined__lte=to_date)
        except ValueError:
            messages.error(request, "Invalid To Date format. Please use '4 March, 2025'.")

    if pick_date:
        try:
            pick_date = datetime.strptime(pick_date, '%d %B, %Y')
            pick_date_start = pick_date.replace(hour=0, minute=0, second=0, microsecond=0)
            pick_date_end = pick_date.replace(hour=23, minute=59, second=59, microsecond=999999)
            users = users.filter(date_joined__range=[pick_date_start, pick_date_end])
        except ValueError:
            messages.error(request, "Invalid Pick Date format. Please use '4 March, 2025'.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Users'

    headers = ['No', 'Username', 'Email', 'Role', 'Active', 'Staff', 'Superuser', 'Date Joined']
    ws.append(headers)

    for index, user in enumerate(users, start=1):
        row = [
            index,
            user.username,
            user.email,
            user.role.name if user.role else None,
            'Yes' if user.is_active else 'No',
            'Yes' if user.is_staff else 'No',
            'Yes' if user.is_superuser else 'No',
            user.date_joined.strftime('%Y-%m-%d %H:%M:%S')
        ]
        ws.append(row)

    current_date = datetime.now().strftime('%Y-%m-%d')
    filename = f'users_{current_date}.xlsx'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response

def UserImportExcel(request):
    if request.method == 'POST' and 'file' in request.FILES:
        excel_file = request.FILES['file']
        
        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            # Iterate over the rows, skipping the header row
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Start from row 2 (skip header)
                
                # Check if the row has 7 columns or more
                if len(row) < 7:
                    messages.error(request, f"Invalid row format. Expected 7 values, but found {len(row)} values. Skipping this row.")
                    continue  # Skip this row if it has fewer than 7 columns
                
                 # Skip the first column (No column) and process the next 7 columns
                row = row[1:]  # Remove the "No" column (first column)
                
                username, email, role_name, is_active, is_staff, is_superuser, date_joined = row

                # Ensure the role exists and is valid
                role, created = RoleModel.objects.get_or_create(name=role_name)

                # Try to get the existing user
                try:
                    user = UserModel.objects.get(email=email)

                    # If user is soft-deleted, restore them
                    if user.deleted_at:
                        user.restore()
                        messages.success(request, f"User {username} restored successfully.")
                    else:
                        messages.info(request, f"User {username} already exists and is active.")
                except UserModel.DoesNotExist:
                    # Create new user if not found
                    try:
                        UserModel.objects.create(
                            username=username,
                            email=email,
                            role=role,
                            is_active=is_active == 'Yes',
                            is_staff=is_staff == 'Yes',
                            is_superuser=is_superuser == 'Yes',
                            date_joined=date_joined
                        )
                        messages.success(request, f"User {username} created successfully.")
                    except IntegrityError:
                        messages.error(request, f"Duplicate entry found for {username}. Skipping this row.")
                        continue 

            messages.success(request, 'Users processed successfully.')
            return redirect('/user/list/')  

        except openpyxl.utils.exceptions.InvalidFileException:
            messages.error(request, 'Invalid file format. Please upload a valid Excel file.')
            return redirect('/user/list/')
        except Exception as e:
            messages.error(request, f'Error occurred while importing: {str(e)}')
            return redirect('/user/list/')
    else:
        messages.error(request, 'No file uploaded.')
        return redirect('/user/list/')

def UserTrashList(request):
    search_term = request.GET.get('search', '')
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')
    pick_date = request.GET.get('pick_date', '')
    page_number = request.GET.get('page', 1)
    entries = request.GET.get('entries', 5)

    try:
        entries = int(entries)
    except ValueError:
        entries = 5
        
    all_users = UserModel.objects.filter(deleted_at = None).order_by('-created_at')
    users = UserModel.objects.filter(deleted_at__isnull=False).order_by('-deleted_at')

    if search_term:
        users = users.filter(
            Q(username__icontains=search_term) |
            Q(email__icontains=search_term) |
            Q(role__name__icontains=search_term)
        )

    if from_date:
        try:
            from_date = datetime.strptime(from_date, '%d %B, %Y')
            users = users.filter(date_joined__gte=from_date)
        except ValueError:
            messages.error(request, "Invalid From Date format. Please use '4 March, 2025'.")

    if to_date:
        try:
            to_date = datetime.strptime(to_date, '%d %B, %Y')
            to_date = to_date.replace(hour=23, minute=59, second=59, microsecond=999999)
            users = users.filter(date_joined__lte=to_date)
        except ValueError:
            messages.error(request, "Invalid To Date format. Please use '4 March, 2025'.")

    if pick_date:
        try:
            pick_date = datetime.strptime(pick_date, '%d %B, %Y')
            pick_date_start = pick_date.replace(hour=0, minute=0, second=0, microsecond=0)
            pick_date_end = pick_date.replace(hour=23, minute=59, second=59, microsecond=999999)
            users = users.filter(date_joined__range=[pick_date_start, pick_date_end])
        except ValueError:
            messages.error(request, "Invalid Pick Date format. Please use '4 March, 2025'.")

    paginator = Paginator(users, entries)
    page_obj = paginator.get_page(page_number)

    roles = RoleModel.objects.all().order_by('-created_at')

    context = {
        "users": page_obj,
        "all_users" : all_users,
        "roles": roles,
        "search": search_term,
        "entries": entries,
        "from_date": from_date,
        "to_date": to_date,
        "pick_date": pick_date,
    }
    return render(request, 'dashboard/user_trash_list.html', context)

@login_required(login_url=settings.LOGIN_URL)
def UserRestore(request, pk):
    user = UserModel.objects.get(id=pk)
    user.restore()
    messages.success(request, "User has been restored!")
    return redirect('/user/trash/list/')
